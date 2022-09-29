from app.model import Collector, Pecuaristas, Answers
from peewee import JOIN
from datetime import datetime, timedelta
import logging
from app.config import config, planilhas_dir
import sys
import os
from openpyxl import Workbook
from openpyxl.styles import Font, numbers
from app.ftp import create_ftp
import requests
from app.email_jbs import EmailJBS
from app.jbs import JBSNotificacao

ftp_pasta_planilha = 'retorno/'
emailj = EmailJBS()
jbs_notificaca = JBSNotificacao()
dict_resps = {'Péssimo': 1,
              'Ruim': 2,
              'Bom': 3,
              'Muito Bom': 4,
              'Excelente': 5,
              '': 0,
              None: None,
              }

dict_counter = {'1': {'Péssimo': 0,
                      'Ruim': 0,
                      'Bom': 0,
                      'Muito Bom': 0,
                      'Excelente': 0,
                      '': -1,
                      },
                '2': {'Péssimo': 0,
                      'Ruim': 0,
                      'Bom': 0,
                      'Muito Bom': 0,
                      'Excelente': 0,
                      '': -1,
                      },
                '3': {'Péssimo': 0,
                      'Ruim': 0,
                      'Bom': 0,
                      'Muito Bom': 0,
                      'Excelente': 0,
                      '': -1,
                      },
                }

STATUS = {'completely_responded': 'Completado', 'partially_responded': 'Parcialmente Respondido', 'overquota': 'overquota', 'disqualified': 'Desqualificado',
          'not_responded': 'Não respondido', 'Não respondeu até a data vigente': 'Não respondeu até a data vigente',
          'sent': 'Enviado', 'not_sent': 'Não Enviado', 'processing': 'Processando', 'bounced': 'Não Entregue', 'opted_out': 'Optou por Sair'
          }

key = config['survey_monkey_key']

f = lambda r, k: r[k] if k in r else k

def status_email(mail_status, response_status):
    if response_status in ['completely_responded', 'partially_responded']:
        return response_status

    if mail_status in ['bounced', 'opted_out']:
        return mail_status

    if mail_status == 'sent':
        return STATUS['Não respondeu até a data vigente']

    return mail_status



days = datetime.today() - timedelta(days=7)

workbook = Workbook()
results_filename = 'PESQUISA_ABATE_JBS.xlsx'
results_file = '{}/{}'.format(planilhas_dir, results_filename)

# open the file created and writes on it
if os.path.exists(results_file):
    os.remove(results_file)

worksheet = workbook.active
worksheet.title = 'Respostas'

if 'VISÃO GERAL' not in workbook.sheetnames:
    workbook.create_sheet('VISÃO GERAL')
sheet_visao_geral = workbook['VISÃO GERAL']


worksheet.cell(row=1, column=1, value="Data do Ultimo Abate").font = Font(bold=True)
worksheet.cell(row=1, column=2, value="Pedido Comercial").font = Font(bold=True)
worksheet.cell(row=1, column=3, value="Codigo Fazenda").font = Font(bold=True)
worksheet.cell(row=1, column=4, value="Pecuarista").font = Font(bold=True)
worksheet.cell(row=1, column=5, value="Sigla").font = Font(bold=True)
worksheet.cell(row=1, column=6, value="Email").font = Font(bold=True)
worksheet.cell(row=1, column=7, value="Telefone").font = Font(bold=True)
worksheet.cell(row=1, column=8, value="Telefone2").font = Font(bold=True)
worksheet.cell(row=1, column=9, value="Celular").font = Font(bold=True)
worksheet.cell(row=1, column=10, value="Status do Email").font = Font(bold=True)
worksheet.cell(row=1, column=11, value="1-Abate").font = Font(bold=True)
worksheet.cell(row=1, column=12, value="2-Atend. Comprador").font = Font(bold=True)
worksheet.cell(row=1, column=13, value="3-Transp. Animais").font = Font(bold=True)
worksheet.cell(row=1, column=14, value="Comentários Resp. 1").font = Font(bold=True)
worksheet.cell(row=1, column=15, value="Comentários Resp. 2").font = Font(bold=True)
worksheet.cell(row=1, column=16, value="Comentários Resp. 3").font = Font(bold=True)
worksheet.cell(row=1, column=17, value="Instrucao").font = Font(bold=True)

worksheet.column_dimensions["A"].width = 15.0
worksheet.column_dimensions["B"].width = 20.0
worksheet.column_dimensions["C"].width = 18.0
worksheet.column_dimensions["D"].width = 50.0
worksheet.column_dimensions["E"].width = 10.0
worksheet.column_dimensions["F"].width = 50.0
worksheet.column_dimensions["G"].width = 15.0
worksheet.column_dimensions["H"].width = 15.0
worksheet.column_dimensions["I"].width = 15.0
worksheet.column_dimensions["J"].width = 30.0
worksheet.column_dimensions["K"].width = 20.0
worksheet.column_dimensions["L"].width = 20.0
worksheet.column_dimensions["M"].width = 20.0
worksheet.column_dimensions["N"].width = 20.0
worksheet.column_dimensions["O"].width = 20.0
worksheet.column_dimensions["P"].width = 20.0
worksheet.column_dimensions["Q"].width = 10.0


logging.info('preparando planilha de respostas')
row = 2
#pecuarista_collection = Pecuaristas.select().where(Pecuaristas.data_ultimo_abate >= datetime(2018, 8, 1))\
#    .join(Answers, JOIN.LEFT_OUTER).order_by(Pecuaristas.data_ultimo_abate.desc(), Answers.resposta_1.desc())
pecuarista_collection = Pecuaristas.select().where(Pecuaristas.data_ultimo_abate >= datetime(2021, 1, 1))\
    .join(Answers, JOIN.LEFT_OUTER).order_by(Pecuaristas.data_ultimo_abate.desc(), Answers.resposta_1.desc())
for pecuarista_model in pecuarista_collection:
    worksheet.cell(row=row, column=1, value=pecuarista_model.data_ultimo_abate).number_format = numbers.FORMAT_DATE_DDMMYY
    worksheet.cell(row=row, column=2, value=pecuarista_model.pedido_comercial)
    worksheet.cell(row=row, column=3, value=pecuarista_model.fazenda)
    worksheet.cell(row=row, column=4, value=pecuarista_model.nome)
    worksheet.cell(row=row, column=5, value=pecuarista_model.empresa)
    worksheet.cell(row=row, column=6, value=pecuarista_model.email)
    worksheet.cell(row=row, column=7, value=pecuarista_model.telefone)
    worksheet.cell(row=row, column=8, value=pecuarista_model.telefone_2)
    worksheet.cell(row=row, column=9, value=pecuarista_model.celular)
    worksheet.cell(row=row, column=17, value=pecuarista_model.instrucao)
    worksheet.cell(row=row, column=10, value='Não Enviado')
    try:
        answer_model = pecuarista_model.answer.get()
        s = f(STATUS, status_email(answer_model.mail_status, answer_model.survey_response_status))
        worksheet.cell(row=row, column=10, value=s if s not in ['', None] else 'Não respondeu até a data vigente')
        worksheet.cell(row=row, column=11, value=answer_model.resposta_1)
        worksheet.cell(row=row, column=12, value=answer_model.resposta_2)
        worksheet.cell(row=row, column=13, value=answer_model.resposta_3)
        worksheet.cell(row=row, column=14, value=answer_model.comentario_1)
        worksheet.cell(row=row, column=15, value=answer_model.comentario_2)
        worksheet.cell(row=row, column=16, value=answer_model.comentario_3)
    except Answers.DoesNotExist:
        pass

    row += 1


#
sheet_visao_geral.append(['DADOS DIÁRIOS', 'Péssimo', 'Ruim', 'Bom', 'Muito Bom', 'Excelente',
                          '', 'Enviados','Não Entregues','Abertos','Não Abertos','Completamente Respondido','Não Respondido','Parcialmente Respondido','Descadastrados','Cliques no Link'])

for c in sheet_visao_geral['A1:P1'][0]:
    c.font = Font(bold=True)
    sheet_visao_geral.column_dimensions[c.column].width = 18

for collector_model in Collector.select().where(Collector.date >= datetime.now() - timedelta(days=60)).order_by(Collector.date.desc()):
    contage_total = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, None: 0}
    logging.info('stats para data {}'.format(collector_model.date))
    for pecuarista_model in Pecuaristas.select().where(Pecuaristas.data_ultimo_abate == (collector_model.date - timedelta(days=1))):
        try:
            answer_model = pecuarista_model.answer.get()
            contage_total[answer_model.resposta_1] += 1
            contage_total[answer_model.resposta_2] += 1
            contage_total[answer_model.resposta_3] += 1
        except Answers.DoesNotExist:
            pass

    stats = requests.get('https://api.surveymonkey.com/v3/collectors/{}/stats'.format(collector_model.collector_id),
                         headers={'Authorization': 'bearer {}'.format(key)}).json()
    row = (sheet_visao_geral.max_row + 1)
    sheet_visao_geral.cell(row=row, column=1, value=(collector_model.date - timedelta(days=1)).strftime('%d/%m/%Y'))

    sheet_visao_geral.cell(row=row, column=2, value=contage_total.get(1))
    sheet_visao_geral.cell(row=row, column=3, value=contage_total.get(2))
    sheet_visao_geral.cell(row=row, column=4, value=contage_total.get(3))
    sheet_visao_geral.cell(row=row, column=5, value=contage_total.get(4))
    sheet_visao_geral.cell(row=row, column=6, value=contage_total.get(5))

    sheet_visao_geral.cell(row=row, column=8, value=stats['mail_status']['sent'])
    sheet_visao_geral.cell(row=row, column=9, value=stats['mail_status']['bounced'])
    sheet_visao_geral.cell(row=row, column=10, value=stats['mail_status']['opened'])
    sheet_visao_geral.cell(row=row, column=11, value=stats['mail_status']['not_sent'])
    sheet_visao_geral.cell(row=row, column=12, value=stats['survey_response_status']['completely_responded'])
    sheet_visao_geral.cell(row=row, column=13, value=stats['survey_response_status']['not_responded'])
    sheet_visao_geral.cell(row=row, column=14, value=stats['survey_response_status']['partially_responded'])
    sheet_visao_geral.cell(row=row, column=15, value=stats['mail_status']['opted_out'])
    sheet_visao_geral.cell(row=row, column=16, value=stats['mail_status']['link_clicked'])


workbook.save(results_file)
workbook.close()
logging.info('planilha finalizada {}'.format(results_file))

#logging.info("uploading {} to FTP Server...".format(results_file))
#ftp = create_ftp()
#ftp.upload(results_file, ftp_pasta_planilha + results_filename)
#logging.info("Uploaded completed successfully")
