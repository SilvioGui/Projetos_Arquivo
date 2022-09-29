from app.jbs import JBSNotificacao
from app.model import JBSNotificacaoModel, SiglasModel, Answers
from app.config import planilhas_dir
from app.email_jbs import EmailJBSManager
import os
from openpyxl import Workbook
from openpyxl.styles import Font, numbers
import logging
import sys

jbs = JBSNotificacao()

STATUS = {'completely_responded': 'Completado', 'partially_responded': 'Parcialmente Respondido', 'overquota': 'overquota', 'disqualified': 'Desqualificado',
          'not_responded': 'Não respondido', 'Não respondeu até a data vigente': 'Não respondeu até a data vigente',
          'sent': 'Enviado', 'not_sent': 'Não Enviado', 'processing': 'Processando', 'bounced': 'Não Entregue', 'opted_out': 'Optou por Sair'
          }

f = lambda r, k: r[k] if k in r else k

def status_email(mail_status, response_status):
    if response_status in ['completely_responded', 'partially_responded']:
        return response_status

    if mail_status in ['bounced', 'opted_out']:
        return mail_status

    if mail_status == 'sent':
        return STATUS['Não respondeu até a data vigente']

    return mail_status

siglas = {}

for notification in jbs.get_all_notification().where(JBSNotificacaoModel.tipo == jbs.TIPO_PROPOSITO_AVALIACAO):
    pecuarista = notification.pecuarista
    if pecuarista.empresa not in siglas:
        siglas[pecuarista.empresa] = []

    siglas[pecuarista.empresa].append(pecuarista)


file = 'Satisfação dos pecuaristas.xlsx'
file_full = '{}/{}'.format(planilhas_dir, file)


destinatariosAbate = 'eduardo.watanabe@friboi.com.br,fabio.dias@friboi.com.br,fernanda.matos@friboi.com.br,daniel.avila@friboi.com.br'
destinatariosAtendimento = 'eduardo.watanabe@friboi.com.br,fabio.dias@friboi.com.br,fernanda.matos@friboi.com.br,ricardo.gelain@jbs.com.br, marcio.salaber@jbs.com.br,leonardo.vieira@friboi.com.br'
destinatariosTransporte = 'eduardo.watanabe@friboi.com.br,fabio.dias@friboi.com.br,fernanda.matos@friboi.com.br,ricardo.gelain@jbs.com.br, marcio.salaber@jbs.com.br,leonardo.vieira@friboi.com.br,klever.vendrame@friboi.com.br'

siglasAbate = [
    {"sigla": "AFT", "dest": [{"email": "thiago.fernando@friboi.com.br", "nome": "THIAGO CAVALLI"}, {"email": "rafael.farah@friboi.com.br", "nome": "RAFAEL FARAH"}]},
    {"sigla": "AGB", "dest": [{"email": "milton.hamada@friboi.com.br", "nome": "MILTON HAMADA"},{"email": "renato.macedo@friboi.com.br", "nome": "RENATO MACEDO"}]},
    {"sigla": "AMS", "dest": [{"email": "marcelo.coelho@friboi.com.br", "nome": "MARCELO LIMA COELHO"},{"email": "marcos.cheida@friboi.com.br", "nome": "MARCOS CHEIDA"}]},
    {"sigla": "ANS", "dest": [{"email": "marcelo.coelho@friboi.com.br", "nome": "MARCELO LIMA COELHO"},{"email": "marcos.cheida@friboi.com.br", "nome": "MARCOS CHEIDA"}]},
    {"sigla": "ANF", "dest": [{"email": "paulo.rinaldi@friboi.com.br", "nome": "PAULO RINALDI"},{"email": "marcos.cheida@friboi.com.br", "nome": "MARCOS CHEIDA"}]},
    {"sigla": "AND", "dest": [{"email": "paulo.rinaldi@friboi.com.br", "nome": "PAULO RINALDI"},{"email": "marcos.cheida@friboi.com.br", "nome": "MARCOS CHEIDA"}]},
    {"sigla": "ANM", "dest": [{"email": "paulo.rinaldi@friboi.com.br", "nome": "PAULO RINALDI"},{"email": "marcos.cheida@friboi.com.br", "nome": "MARCOS CHEIDA"}]},
    {"sigla": "ARA", "dest": [{"email": "gilmar.alves@friboi.com.br", "nome": "GILMAR ALVES DE JESUS"}, {"email": "rafael.farah@friboi.com.br", "nome": "RAFAEL FARAH"}]},
    {"sigla": "ATO", "dest": [{"email": "bruno.brainer@friboi.com.br", "nome": "BRUNO BRAINER"}]},
    {"sigla": "BAR", "dest": [{"email": "ivan.ritter@friboi.com.br", "nome": "IVAN ANTONIO RITTER"},{"email": "renato.macedo@friboi.com.br", "nome": "RENATO MACEDO"}]},
    {"sigla": "CFS", "dest": [{"email": "laercio.guagliano@friboi.com.br", "nome": "LAERCIO GUAGLIANO"},{"email": "renato.macedo@friboi.com.br", "nome": "RENATO MACEDO"}]},
    {"sigla": "CGR", "dest": [{"email": "jose.piva@friboi.com.br", "nome": "JOSE VITORIO PIVA"},{"email": "marcos.cheida@friboi.com.br", "nome": "MARCOS CHEIDA"}]},
    {"sigla": "CLR", "dest": [{"email": "rubens.bigaran@friboi.com.br", "nome": "RUBENS BIGARAN DA SILVA"},{"email": "rafael.farah@friboi.com.br", "nome": "RAFAEL FARAH"}]},
    {"sigla": "CPG", "dest": [{"email": "aloisio.mastelaro@friboi.com.br", "nome": "ALOISIO ANTONIO MASTELARO JUNIOR"},{"email": "marcos.cheida@friboi.com.br", "nome": "MARCOS CHEIDA"}]},
    {"sigla": "CXI", "dest": [{"email": "luiz.braga@friboi.com.br", "nome": "LUIZ CARNEIRO BRAGA"},{"email": "marcos.cheida@friboi.com.br", "nome": "MARCOS CHEIDA"}]},
    {"sigla": "DMT", "dest": [{"email": "osmar.junior@friboi.com.br", "nome": "OSMAR DE PADUA CARNEIRO JUNIOR"},{"email": "rafael.farah@friboi.com.br", "nome": "RAFAEL FARAH"}]},
    {"sigla": "GYN", "dest": [{"email": "lidiane.surpilli@friboi.com.br", "nome": "LIDIANE SURPILLI"},{"email": "renato.macedo@friboi.com.br", "nome": "RENATO MACEDO"}]},
    {"sigla": "ITA", "dest": [{"email": "deyvdson.lima@friboi.com.br", "nome": "DEYVDSON DA CONCEICAO DE LIMA"}]},
    {"sigla": "ITB", "dest": [{"email": "lucas.chaves@friboi.com.br", "nome": "LUCAS CHAVES DE FIGUEIREDO"},{"email": "renato.macedo@friboi.com.br", "nome": "RENATO MACEDO"}]},
    {"sigla": "ITR", "dest": [{"email": "mauro.serpa@friboi.com.br", "nome": "MAURO SERPA"},{"email": "renato.macedo@friboi.com.br", "nome": "RENATO MACEDO"}]},
    {"sigla": "JUA", "dest": [{"email": "alisson.lima@friboi.com.br", "nome": "ALISSON DE SOUZA LIMA"},{"email": "rafael.farah@friboi.com.br", "nome": "RAFAEL FARAH"}]},
    {"sigla": "JUI", "dest": [{"email": "reinaldo.morais@friboi.com.br", "nome": "REINALDO MIRANDA MORAIS"},{"email": "rafael.farah@friboi.com.br", "nome": "RAFAEL FARAH"}]},
    {"sigla": "LIF", "dest": [{"email": "toni.peixoto@friboi.com.br", "nome": "TONI DUARTE PEIXOTO"},{"email": "marcos.cheida@friboi.com.br", "nome": "MARCOS CHEIDA"}]},
    {"sigla": "LIA", "dest": [{"email": "toni.peixoto@friboi.com.br", "nome": "TONI DUARTE PEIXOTO"},{"email": "marcos.cheida@friboi.com.br", "nome": "MARCOS CHEIDA"}]},
    {"sigla": "LIN", "dest": [{"email": "toni.peixoto@friboi.com.br", "nome": "TONI DUARTE PEIXOTO"},{"email": "marcos.cheida@friboi.com.br", "nome": "MARCOS CHEIDA"}]},
    {"sigla": "MRB", "dest": [{"email": "ricardo.pires@friboi.com.br", "nome": "RICARDO PIRES"},{"email": "bruno.brainer@friboi.com.br", "nome": "BRUNO BRAINER"}]},
    {"sigla": "MZL", "dest": [{"email": "edson.raymundi@friboi.com.br", "nome": "EDSON RAYMUNDI"},{"email": "renato.macedo@friboi.com.br", "nome": "RENATO MACEDO"}]},
    {"sigla": "NAD", "dest": [{"email": "luiz.pacheco@friboi.com.br", "nome": "LUIZ PACHECO"},{"email": "marcos.cheida@friboi.com.br", "nome": "MARCOS CHEIDA"}]},
    {"sigla": "NVR", "dest": [{"email": "fernando.zulato@friboi.com.br", "nome": "LUIS FERNANDO ZULATO"},{"email": "marcos.cheida@friboi.com.br", "nome": "MARCOS CHEIDA"}]},
    {"sigla": "PDP", "dest": [{"email": "carlos.timoteo@friboi.com.br", "nome": "CARLOS SERGIO TIMOTEO"},{"email": "renato.macedo@friboi.com.br", "nome": "RENATO MACEDO"}]},
    {"sigla": "PEL", "dest": [{"email": "waldir.senhor@friboi.com.br", "nome": "WALDIR DA SILVA SENHOR"},{"email": "rafael.farah@friboi.com.br", "nome": "RAFAEL FARAH"}]},
    {"sigla": "PIB", "dest": [{"email": "vinicius.paro@friboi.com.br", "nome": "VINICIUS DOMINGOS PARO"},{"email": "rafael.farah@friboi.com.br", "nome": "RAFAEL FARAH"}]},
    {"sigla": "POP", "dest": [{"email": "luiz.ciriaco@friboi.com.br", "nome": "LUIZ DAVID CIRIACO"},{"email": "marcos.cheida@friboi.com.br", "nome": "MARCOS CHEIDA"}]},
    {"sigla": "PVH", "dest": [{"email": "arnaldo.ramalho@friboi.com.br", "nome": "ARNALDO RAMALHO DE CALDAS"}]},
    {"sigla": "RBR", "dest": [{"email": "hudson.junior@friboi.com.br", "nome": "HUDSON JUNIOR"},{"email":"arnaldo.ramalho@friboi.com.br", "nome": "ARNALDO RAMALHO DE CALDAS"}]},
    {"sigla": "RED", "dest": [{"email": "bruno.brainer@friboi.com.br", "nome": "BRUNO BRAINER"}]},
    {"sigla": "SEN", "dest": [{"email": "henrique.haddad@friboi.com.br", "nome": "HENRIQUE PASSONI HADDAD"},{"email": "renato.macedo@friboi.com.br", "nome": "RENATO MACEDO"}]},
    {"sigla": "SMG", "dest": [{"email": "aecio.pereira@friboi.com.br", "nome": "AECIO WIRLEY PEREIRA"},{"email": "rafael.farah@friboi.com.br", "nome": "RAFAEL FARAH"}]},
    {"sigla": "STA", "dest": [{"email": "rafael.goulding@friboi.com.br", "nome": "RAFAEL GOULDING"},{"email": "bruno.brainer@friboi.com.br", "nome": "BRUNO BRAINER"}]},
    {"sigla": "TCM", "dest": [{"email": "marlon.santos@friboi.com.br", "nome": "MARLON DENNY SANTOS"},{"email": "bruno.brainer@friboi.com.br", "nome": "BRUNO BRAINER"}]},
    {"sigla": "VHA", "dest": [{"email": "lessandro.caprini@friboi.com.br", "nome": "LESSANDRO MENIN CAPRINI"},{"email": "rafael.farah@friboi.com.br", "nome": "RAFAEL FARAH"}]}
]

for sigla, items in siglas.items():
    resp1 = []
    resp2 = []
    resp3 = []
    logging.info('sigla {} quantidade de itens {}'.format(sigla, len(items)))
    if os.path.exists(file_full):
        os.remove(file_full)

    workbook = Workbook()
    ws = workbook.active
    ws.title = 'Respostas'

    ws.cell(row=1, column=1, value="Data do Ultimo Abate").font = Font(bold=True)
    ws.cell(row=1, column=2, value="Pedido Comercial").font = Font(bold=True)
    ws.cell(row=1, column=3, value="Codigo Fazenda").font = Font(bold=True)
    ws.cell(row=1, column=4, value="Pecuarista").font = Font(bold=True)
    ws.cell(row=1, column=5, value="Sigla").font = Font(bold=True)
    ws.cell(row=1, column=6, value="Email").font = Font(bold=True)
    ws.cell(row=1, column=7, value="Telefone").font = Font(bold=True)
    ws.cell(row=1, column=8, value="Telefone2").font = Font(bold=True)
    ws.cell(row=1, column=9, value="Celular").font = Font(bold=True)
    ws.cell(row=1, column=10, value="Status do Email").font = Font(bold=True)
    ws.cell(row=1, column=11, value="1-Abate").font = Font(bold=True)
    ws.cell(row=1, column=12, value="2-Atend. Comprador").font = Font(bold=True)
    ws.cell(row=1, column=13, value="3-Transp. Animais").font = Font(bold=True)
    ws.cell(row=1, column=14, value="Comentários Resp. 1").font = Font(bold=True)
    ws.cell(row=1, column=15, value="Comentários Resp. 2").font = Font(bold=True)
    ws.cell(row=1, column=16, value="Comentários Resp. 3").font = Font(bold=True)
    ws.cell(row=1, column=17, value="Instrução").font = Font(bold=True)

    ws.column_dimensions["A"].width = 15.0
    ws.column_dimensions["B"].width = 20.0
    ws.column_dimensions["C"].width = 18.0
    ws.column_dimensions["D"].width = 50.0
    ws.column_dimensions["E"].width = 10.0
    ws.column_dimensions["F"].width = 50.0
    ws.column_dimensions["G"].width = 15.0
    ws.column_dimensions["H"].width = 15.0
    ws.column_dimensions["I"].width = 15.0
    ws.column_dimensions["J"].width = 30.0
    ws.column_dimensions["K"].width = 20.0
    ws.column_dimensions["L"].width = 20.0
    ws.column_dimensions["M"].width = 20.0
    ws.column_dimensions["N"].width = 20.0
    ws.column_dimensions["O"].width = 20.0
    ws.column_dimensions["P"].width = 20.0
    ws.column_dimensions["Q"].width = 10.0

    row = 2
    for pecuarista_model in items:

        if pecuarista_model.answer.get().resposta_1 in [1, 2]:
            resp1.append(pecuarista_model)
        if pecuarista_model.answer.get().resposta_2 in [1, 2]:
            resp2.append(pecuarista_model)
        if pecuarista_model.answer.get().resposta_3 in [1, 2]:
            resp3.append(pecuarista_model)

        ws.cell(row=row, column=1,
                       value=pecuarista_model.data_ultimo_abate).number_format = numbers.FORMAT_DATE_DDMMYY
        ws.cell(row=row, column=2, value=pecuarista_model.pedido_comercial)
        ws.cell(row=row, column=3, value=pecuarista_model.fazenda)
        ws.cell(row=row, column=4, value=pecuarista_model.nome)
        ws.cell(row=row, column=5, value=pecuarista_model.empresa)
        ws.cell(row=row, column=6, value=pecuarista_model.email)
        ws.cell(row=row, column=7, value=pecuarista_model.telefone)
        ws.cell(row=row, column=8, value=pecuarista_model.telefone_2)
        ws.cell(row=row, column=9, value=pecuarista_model.celular)
        ws.cell(row=row, column=17, value=pecuarista_model.instrucao)

        try:
            answer_model = pecuarista_model.answer.get()
            ws.cell(row=row, column=10,
                           value=f(STATUS, status_email(answer_model.mail_status, answer_model.survey_response_status)))
            ws.cell(row=row, column=11, value=answer_model.resposta_1)
            ws.cell(row=row, column=12, value=answer_model.resposta_2)
            ws.cell(row=row, column=13, value=answer_model.resposta_3)
            ws.cell(row=row, column=14, value=answer_model.comentario_1)
            ws.cell(row=row, column=15, value=answer_model.comentario_2)
            ws.cell(row=row, column=16, value=answer_model.comentario_3)
        except Answers.DoesNotExist:
            pass

        n = pecuarista_model.jbs_notificacao.get()
        n.enviado = True
        n.save()
        row += 1

    # enviando
    email = EmailJBSManager()
    sigla_model = SiglasModel.get(SiglasModel.sigla.contains(sigla))

    workbook.save(file_full)
    to = []
    if sigla_model.gerente:
        to = to + sigla_model.gerente.split(',')
    if sigla_model.gerente_regional:
        to = to + sigla_model.gerente_regional.split(',')

    if len(resp1) > 0:
        for dest in destinatariosAbate.split(','):
            if dest not in to:
                to.append(dest)

        for dest in siglasAbate:
            if sigla == dest['sigla']:
                [to.append(s['email']) for s in dest['dest']]
    if len(resp2) > 0:
        for dest in destinatariosAtendimento.split(','):
            if dest not in to:
                to.append(dest)
    if len(resp3) > 0:
        for dest in destinatariosTransporte.split(','):
            if dest not in to:
                to.append(dest)

    if len(to) > 0:
        logging.info('enviando')
        email.send(file_full, to, sigla_nome=sigla_model.nome)
