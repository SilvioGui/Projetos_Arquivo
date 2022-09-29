from app.model import JBSNotificacaoModel, SiglasModel, UsuariosAppModel, Answers, Pecuaristas
from app.jbs import JBSNotificacao
from openpyxl import Workbook
from openpyxl.styles import Font, numbers
from app.utils import interval_date_without_holiday
from datetime import timedelta
import logging
from app.ftp import create_ftp
from app.config import planilhas_dir
import sys
from datetime import datetime

FILENAME = 'log_app.xlsx'

answer = JBSNotificacao.Answer()
workbook = Workbook()
ws = workbook.active
ws.title = 'app_log'

ws.cell(row=1, column=1, value="Cd Agenda").font = Font(bold=True)
ws.cell(row=1, column=2, value="Descricao").font = Font(bold=True)
ws.cell(row=1, column=3, value="Comentário").font = Font(bold=True)
ws.cell(row=1, column=4, value="Cd Fazenda").font = Font(bold=True)
ws.cell(row=1, column=5, value="Cd Pecuarista").font = Font(bold=True)
ws.cell(row=1, column=6, value="Prazo").font = Font(bold=True)
ws.cell(row=1, column=7, value="Email Usuario").font = Font(bold=True)
ws.cell(row=1, column=8, value="Proposito").font = Font(bold=True)

ws.column_dimensions["A"].width = 15.0
ws.column_dimensions["B"].width = 45.0
ws.column_dimensions["C"].width = 25.0
ws.column_dimensions["D"].width = 12.0
ws.column_dimensions["E"].width = 15.0
ws.column_dimensions["F"].width = 30.0
ws.column_dimensions["G"].width = 25.0

if JBSNotificacaoModel.select().count() == 0:
    logging.info('nao existe registros na tabela de notificacao')
    sys.exit()

logging.info('iniciando processo com {} registros'.format(JBSNotificacaoModel.select().count()))
for n in JBSNotificacaoModel.select().order_by(JBSNotificacaoModel.created.desc()):
    pecuarista_model = n.pecuarista
    try:
        siglas_model = SiglasModel.get(SiglasModel.sigla.contains(pecuarista_model.empresa))
    except Exception:
        continue

    try:
        usuarios_app_mode = UsuariosAppModel.get(email=siglas_model.responsavel)
    except UsuariosAppModel.DoesNotExist:
        logging.info('sem match {}'.format(siglas_model.responsavel))
        continue

    if n.tipo == JBSNotificacao.TIPO_PROPOSITO_EMAIL_VAZIO:
        cnt = Pecuaristas.select().where(
            Pecuaristas.created <= pecuarista_model.created,
            Pecuaristas.created >= interval_date_without_holiday(pecuarista_model.created, timedelta(days=-2)),
            Pecuaristas.codigo_pecuarista == pecuarista_model.codigo_pecuarista,
            Pecuaristas.id != n.pecuarista
        ).count()

        # nao enviar pecuarista repetido para ajuste de cadastro
        if cnt > 0:
            continue

        ws.append(['', JBSNotificacao.DESCRICAO_SEM_EMAIL, '', pecuarista_model.fazenda, pecuarista_model.codigo_pecuarista,
                   interval_date_without_holiday(n.created, timedelta(days=2)), usuarios_app_mode.email,  n.tipo])

    if n.tipo == JBSNotificacao.TIPO_PROPOSITO_AVALIACAO:
        try:
            answer_model = pecuarista_model.answer.get()
            for a in answer.parse_resposta(answer_model):
                ws.append(['', a[0], a[1], pecuarista_model.fazenda, pecuarista_model.codigo_pecuarista, interval_date_without_holiday(n.created, timedelta(days=2)), usuarios_app_mode.email, n.tipo])

        except Answers.DoesNotExist:
            pass
for c in ws['E']:
    c.number_format = numbers.FORMAT_DATE_DDMMYY

workbook.save('{}/{}'.format(planilhas_dir, FILENAME))
logging.info('enviando arquivo para ftp')
f = create_ftp()
f.upload('{}/{}'.format(planilhas_dir, FILENAME), 'retorno/'+FILENAME)
logging.info('finalizado com sucesso')
